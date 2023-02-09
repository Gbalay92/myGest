import openpyxl as op
import shutil as sh
from datetime import date as dt
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

def crear_nominas():
    original="Documentos/plantillas/nominas/nomina-11223344.xlsx"
    wb=op.load_workbook("Documentos/rrhh/empregados.xlsx")
    categorias=wb['taboas Salariais']
    empleados=wb['empregados']

    dict_categorias={}
    for row in categorias.iter_rows(min_row=2, max_col=2):
        dict_categorias[row[0].value] = row[1].value
        
    for row in empleados.iter_rows(min_row=2, max_col=4):
        folder="Documentos/rrhh/nominas/"+str(datetime.now().month)+"-"+str(datetime.now().year)+"/"
        os.makedirs(os.path.dirname(folder), exist_ok=True)
        fichero=folder+"/"+ row[0].value.replace(" ", "")+str(datetime.now().month)+"-"+str(datetime.now().year)+".xlsx"
        sh.copy(original, fichero)
        wb=op.load_workbook(fichero)
        ws=wb.active
        ws['C8']=str(dt.today().replace(day=1).strftime("%d/%m/%Y"))
        ws['D8']=str((dt.today() + relativedelta(day=31)).strftime("%d/%m/%Y"))
        ws['E2']=row[0].value
        ws['F5']=row[1].value
        ws['B38']=row[2].value
        ws['F15']=row[3].value
        ws['F11']=dict_categorias[row[1].value]
        wb.save(fichero)

