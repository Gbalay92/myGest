from modulos.models.Transaccion import Compra, Venta
import openpyxl as op
import os
import shutil as sh
import formulas
from datetime import datetime, timedelta, date as dt
from dateutil.relativedelta import relativedelta


dict_cell_gerencia = {1:"D",
                    2:"E",
                    3:"F",
                    4:"G",
                    5:"H",
                    6:"I",
                    7:"J",
                    8:"K",
                    9:"L",
                    10:"M",
                    11:"N",
                    12:"O",}

dict_cell_iva =  {1:"C",
                    2:"D",
                    3:"E",
                    4:"G",
                    5:"H",
                    6:"I",
                    7:"K",
                    8:"L",
                    9:"M",
                    10:"O",
                    11:"P",
                    12:"Q",}

def get_cell_value(solution, file, sheet, cell):
    cell_ref=f"'[{file}]{sheet}'!{cell}"
    return solution.get(cell_ref).values[cell_ref]

def xldate_to_datetime(xldate):
	temp = datetime(1899, 12, 30)
	delta = timedelta(days=xldate)
	return temp+delta


def create_payroll():
    original="Documentos/nominas/nomina-11223344.xlsx"
    wb=op.load_workbook("Documentos/rrhh/empregados.xlsx")
    categorias=wb['taboas Salariais']
    empleados=wb['empregados']
    dict_categorias={}
    for row in categorias.iter_rows(min_row=2, max_col=2):
        dict_categorias[row[0].value] = row[1].value    
    for row in empleados.iter_rows(min_row=2, max_col=4):
        folder="Documentos/nominas/"+str(datetime.now().month)+"-"+str(datetime.now().year)+"/"
        os.makedirs(os.path.dirname(folder), exist_ok=True)
        fichero=folder+"/"+ row[0].value.replace(" ", "")+str(datetime.now().month)+"-"+str(datetime.now().year)+".xlsx"
        sh.copy(original, fichero)
        wb=op.load_workbook(fichero)
        ws=wb.active
        ws['C8']=str(dt.today().replace(day=1).strftime("%d/%m/%Y"))
        ws['D8']=str((dt.today() + relativedelta(day=31)).strftime("%d/%m/%Y"))
        ws['E2']=row[0].value
        ws['F5']=row[1].value
        ws['B38']=row[2].value
        ws['F15']=row[3].value
        ws['F11']=dict_categorias[row[1].value]
        wb.save(fichero)
    
    
def crear_objetos_compra():
    lista_compras = []
    path="Documentos/facturas/compras"
    filename="compras.xlsx"
    xl_model=formulas.ExcelModel().loads(os.path.join(path,filename)).finish()
    solution=xl_model.calculate()
    row=3
    for i in solution:
        try:
            fecha=get_cell_value(solution,filename,"T3-2022","A"+str(row))
            total=get_cell_value(solution,filename,"T3-2022","E"+str(row))
            iva=get_cell_value(solution,filename,"T3-2022","F"+str(row))
            for f in fecha:
                if str(type(f))=="<class 'numpy.ndarray'>":
                    lista_compras.append(Compra(xldate_to_datetime(f[0][0]).strftime("%d/%m/%Y"), iva[-1][0][0], total[-1][0][0]))
            row+=1
        except AttributeError:
            break
    return lista_compras
        
def crear_objetos_venta():
    lista_ventas = []
    dir_path="Documentos/facturas/ventas"
    for file in os.listdir(dir_path): 
        xl_model=formulas.ExcelModel().loads(os.path.join(dir_path,file)).finish()
        solution=xl_model.calculate()
        total=get_cell_value(solution,file,"FACTURA","F47")
        fecha=get_cell_value(solution,file,"FACTURA","H3")
        iva=get_cell_value(solution,file,"FACTURA","F48")
        gasto_mes=total[-1][0][0]
        f=fecha[-1][0][0]
        i=iva[-1][0][0]
        lista_ventas.append(Venta(xldate_to_datetime(f).strftime("%d/%m/%Y") , i, gasto_mes))
        
        #lista_ventas.append(Venta(fecha, iva, total))
    return lista_ventas

def process_transactions(nombre_fichero,header1,header2,header3,lista_a_procesar):
    filepath = "Documentos/facturas/"+nombre_fichero+".xlsx"
    wb = op.Workbook()
    ws=wb.active
    ws['A1']=header1
    ws['B1']=header2
    ws['C1']=header3
    for element in lista_a_procesar:
        ws.append([element.fecha,element.iva,element.total_sin_iva])
    wb.save(filepath)

def write_reports(cantidad, celda, filepath):
    wb = op.load_workbook(filepath)
    ws=wb.active
    ws[f'{celda}']=cantidad
    wb.save(filepath)

def process_gasto_nomina(dict):
    directory="Documentos/nominas"
    for folder in os.listdir(directory):
        dir_path = os.path.join(directory,folder)
        if(dir_path.endswith("xlsx")):
            return
        gasto_mes=0
        mes=folder.split("-")[0]
        for path in os.listdir(dir_path): 
            xl_model=formulas.ExcelModel().loads(os.path.join(dir_path,path)).finish()
            solution=xl_model.calculate()
            values=get_cell_value(solution,path,"NOMINA2","F54")
            gasto_mes+=values[-1][0][0]
        if(len(os.listdir(dir_path))>0):
            write_reports(gasto_mes,dict.get(int(mes))+str(3), "Documentos/gerencia/informe-xerencia.xlsx")
            
def process_report(file_origin, row_destiny,row_orign, file_destiny, dict):
    wb=op.load_workbook(file_origin)
    ws=wb.active
    previous_date=None
    total=0
    for row in ws.iter_rows(min_row=2):
        total+=row[row_orign].value
        date=datetime.strptime(row[0].value,"%d/%m/%Y").month
        if date != previous_date and previous_date != None:
            write_reports(total, dict.get(date)+str(4), file_destiny)
        previous_date=date
    write_reports(total, dict.get(date)+str(row_destiny), file_destiny)
        